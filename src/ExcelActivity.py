from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


class ExcelActivity:
    _currencyRUB = NamedStyle(name="currencyUSD",
                              number_format='_(₽* #,##0.00_);_(₽* (#,##0.00);_(₽* " - "??_);_(@_)')

    def __init__(self, filepath: str):
        self.filename = filepath
        self.workbook = Workbook()
        self.sheet = self.workbook.active

    def write_dataframe(self, df: pd.DataFrame, start_col: int = 1, sheet_name: str = 'Лист1') -> None:
        """
        Записывает таблицу в эксель
        :param df: Таблица DataFrame
        :param start_col: Выбор первого столбца вставки
        :param sheet_name: Название листа Excel, куда необходимо вставить таблицу
        """
        self.sheet.title = sheet_name
        for i, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for j, value in enumerate(row, start=start_col):
                self.sheet.cell(row=i, column=j, value=value)

    def save(self) -> None:
        """
        Сохраняет файл Excel
        """
        self.workbook.save(self.filename)

    def close(self) -> None:
        """
        Закрывает файл Excel
        """
        self.workbook.close()

    def autofit(self) -> None:
        """
        Автоширина ячеек в файле Excel
        """
        dims = {}
        for row in self.sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) * 1.2))
        for col, value in dims.items():
            self.sheet.column_dimensions[col].width = value

    def change_number_format(self, col: str) -> None:
        """
        Изменяет стиль ячеек Excel на финансовый
        :param col: Столбец, в котором необходимо изменить стиль
        """
        target_style = self._currencyRUB
        for cell in self.sheet[col]:
            cell.style = target_style

    def get_rows(self) -> int:
        """
        Возвращает количество строк в файле
        :return: Количество строк в файле
        """
        return self.sheet.max_row

    def add_formulae(self, col: str) -> None:
        """
        Добавляет формулу деления и формулу на сумму в Excel
        :param col: Столбец формулы деления
        """
        self.sheet[f'{col}1'].value = "Результат"
        for row in range(2, self.get_rows() + 1):
            cell = self.sheet[f'{col}{row}']
            cell.number_format = '#,##0.00'
            cell.value = f'=IFERROR(B{row}/E{row}, 0)'

        max_row = self.get_rows()
        self.sheet[f"B{max_row + 1}"].value = f"=SUM(B2:B{max_row})"
        self.sheet[f"E{max_row + 1}"].value = f"=SUM(E2:E{max_row})"
